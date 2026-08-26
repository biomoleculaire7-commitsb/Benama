import os
import json
import time
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, render_template
import db as db
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Search several likely locations for the data file, since manual GitHub
# uploads sometimes drop files at repo root instead of inside data/.
_CANDIDATE_PATHS = [
    os.path.join(BASE_DIR, "data", "school_data.json"),
    os.path.join(BASE_DIR, "school_data.json"),
    os.path.join(BASE_DIR, "..", "school_data.json"),
    os.path.join(BASE_DIR, "..", "data", "school_data.json"),
]
DATA_FILE = next((p for p in _CANDIDATE_PATHS if os.path.exists(p)), None)
if DATA_FILE is None:
    raise FileNotFoundError(
        "تعذّر إيجاد school_data.json. تأكد من رفعه إلى المستودع "
        "(داخل مجلد data/ أو في الجذر مباشرة). تم البحث في: "
        + ", ".join(_CANDIDATE_PATHS)
    )

DOCS_FILE = os.path.join(os.path.dirname(DATA_FILE), "documents.json")
ANN_FILE = os.path.join(os.path.dirname(DATA_FILE), "announcements.json")
ABS_FILE = os.path.join(os.path.dirname(DATA_FILE), "absences.json")
STAFF_ABS_FILE = os.path.join(os.path.dirname(DATA_FILE), "staff_absences.json")
PARENT_PHONES_FILE = os.path.join(os.path.dirname(DATA_FILE), "parent_phones.json")
GUID_FILE = os.path.join(os.path.dirname(DATA_FILE), "guidance.json")
LOGIN_LOG_FILE = os.path.join(os.path.dirname(DATA_FILE), "login_log.json")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
MAX_FILE_SIZE = 15 * 1024 * 1024  # 15MB - real server, no artifact-storage limit

os.makedirs(UPLOAD_DIR, exist_ok=True)

# Same resilience as the data file: find templates/ wherever it actually landed.
_TEMPLATE_CANDIDATES = [
    os.path.join(BASE_DIR, "templates"),
    BASE_DIR,
]
TEMPLATE_DIR = next(
    (p for p in _TEMPLATE_CANDIDATES if os.path.exists(os.path.join(p, "index.html"))),
    os.path.join(BASE_DIR, "templates"),
)

app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE

with open(DATA_FILE, encoding="utf-8") as f:
    SCHOOL_DATA = json.load(f)

STUDENTS = SCHOOL_DATA["students"]
STAFF = SCHOOL_DATA["staff"]
ASSIGNMENTS = SCHOOL_DATA["assignments"]

try:
    db.init_db()  # no-op if DATABASE_URL isn't set — JSON files keep working meanwhile
except Exception as e:
    # A misconfigured DATABASE_URL must NEVER take the whole platform down.
    # Fall back to local JSON storage and keep serving the site.
    print(f"[WARNING] Database connection failed, falling back to JSON files: {e}")
    db.DB_ENABLED = False


def load_docs():
    if os.path.exists(DOCS_FILE):
        with open(DOCS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_docs(docs):
    with open(DOCS_FILE, "w", encoding="utf-8") as f:
        json.dump(docs, f, ensure_ascii=False, indent=2)


def load_json(path, default):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def today_str():
    return datetime.now().strftime("%Y-%m-%d")


def log_login(role, uid, name):
    """Record every login (student/teacher/admin) for the director's daily report."""
    if db.DB_ENABLED:
        db.log_login(role, uid, name)
        return
    logs = load_json(LOGIN_LOG_FILE, [])
    logs.append({
        "role": role,
        "id": uid,
        "name": name,
        "date": today_str(),
        "time": datetime.now().strftime("%H:%M:%S"),
    })
    # keep the log from growing forever on the free tier's small disk
    logs = logs[-3000:]
    save_json(LOGIN_LOG_FILE, logs)


def classify_admin_role(role_text):
    role_text = role_text or ""
    if "مدير" in role_text and "مساعد" not in role_text:
        return "director"
    if "ناظر" in role_text:
        return "supervisor"
    if "مستشار" in role_text and "توجيه" in role_text:
        return "counselor"
    return None


# ---------------- Pages ----------------

@app.route("/")
def index():
    return render_template("index.html")


# ---------------- Auth ----------------

@app.route("/api/login/student", methods=["POST"])
def login_student():
    data = request.get_json(force=True)
    nid = str(data.get("national_id", "")).strip()
    if not nid:
        return jsonify({"ok": False, "error": "يرجى إدخال الرقم التعريف الوطني."}), 400
    stu = next((s for s in STUDENTS if str(s["national_id"]) == nid), None)
    if not stu:
        return jsonify({"ok": False, "error": "هذا الرقم غير مسجّل في قاعدة بياناتنا الحالية."}), 404
    log_login("student", nid, f"{stu['last_name']} {stu['first_name']} ({stu['class']})")
    return jsonify({"ok": True, "user": {**stu, "role": "student"}})


@app.route("/api/login/teacher", methods=["POST"])
def login_teacher():
    data = request.get_json(force=True)
    emp_id = str(data.get("employee_id", "")).strip()
    if not emp_id:
        return jsonify({"ok": False, "error": "يرجى إدخال الرقم التعريف الوظيفي."}), 400
    staff_row = next((s for s in STAFF if str(s["employee_id"]) == emp_id), None)
    if not staff_row:
        return jsonify({"ok": False, "error": "هذا الرقم الوظيفي غير مسجّل في قاعدة بياناتنا الحالية."}), 404
    my_classes = [a for a in ASSIGNMENTS if str(a["employee_id"]) == emp_id]
    if not my_classes:
        return jsonify({
            "ok": False,
            "error": "هذا الرقم مسجّل ضمن الطاقم، لكن لا يوجد له إسناد تربوي حالياً (قد يكون إدارياً)."
        }), 404
    log_login("teacher", emp_id, f"{staff_row['last_name']} {staff_row['first_name']} ({staff_row['subject']})")
    user = {**staff_row, "role": "teacher", "classes": my_classes}
    return jsonify({"ok": True, "user": user})


@app.route("/api/login/admin", methods=["POST"])
def login_admin():
    data = request.get_json(force=True)
    emp_id = str(data.get("employee_id", "")).strip()
    if not emp_id:
        return jsonify({"ok": False, "error": "يرجى إدخال الرقم التعريف الوظيفي."}), 400
    staff_row = next((s for s in STAFF if str(s["employee_id"]) == emp_id), None)
    if not staff_row:
        return jsonify({"ok": False, "error": "هذا الرقم الوظيفي غير مسجّل في قاعدة بياناتنا الحالية."}), 404
    admin_role = classify_admin_role(staff_row.get("role", ""))
    if not admin_role:
        return jsonify({
            "ok": False,
            "error": "هذا الرقم مسجّل ضمن الطاقم، لكنه ليس من فئة الإدارة (مدير / ناظر / مستشار توجيه)."
        }), 404
    log_login("admin", emp_id, f"{staff_row['last_name']} {staff_row['first_name']} ({staff_row['role']})")
    user = {**staff_row, "role": "admin", "admin_role": admin_role}
    return jsonify({"ok": True, "user": user})


# ---------------- Class rosters ----------------

@app.route("/api/roster/<class_name>")
def roster(class_name):
    rows = [dict(s) for s in STUDENTS if s["class"] == class_name]
    phones = db.get_parent_phones_bulk() if db.DB_ENABLED else load_json(PARENT_PHONES_FILE, {})
    for r in rows:
        r["parent_phone"] = phones.get(str(r["national_id"]), "")
    return jsonify({"class": class_name, "count": len(rows), "students": rows})


@app.route("/api/parent-phone/<national_id>", methods=["POST"])
def set_parent_phone(national_id):
    data = request.get_json(force=True)
    phone = (data.get("phone") or "").strip()
    if db.DB_ENABLED:
        db.set_parent_phone(str(national_id), phone)
    else:
        phones = load_json(PARENT_PHONES_FILE, {})
        phones[str(national_id)] = phone
        save_json(PARENT_PHONES_FILE, phones)
    return jsonify({"ok": True})


@app.route("/api/roster/<class_name>/export.xlsx")
def export_roster_csv(class_name):
    rows = [s for s in STUDENTS if s["class"] == class_name]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import Response
    from urllib.parse import quote
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = class_name
    ws.sheet_view.rightToLeft = True

    GREEN = "0F4C3A"
    OCHRE = "C17817"
    thin = Side(style="thin", color="D8CFB8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncols = 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = f"متوسطة الشهيد بن نعمة مصطفى — قائمة تلاميذ قسم {class_name}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=GREEN)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["الرقم", "اللقب", "الاسم", "الجنس", "الرقم التعريف الوطني"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=OCHRE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for idx, s in enumerate(rows, start=1):
        r = idx + 3
        values = [idx, s["last_name"], s["first_name"], s["gender"], s["national_id"]]
        for c_i, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c_i, value=val)
            cell.font = Font(name="Arial", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FBFAF6")

    widths = [7, 20, 20, 8, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"قائمة_قسم_{class_name}.xlsx"
    encoded_filename = quote(filename)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=roster.xlsx; filename*=UTF-8''{encoded_filename}",
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        }
    )


# ---------------- Documents (send/receive) ----------------

@app.route("/api/docs/<class_name>", methods=["GET"])
def get_docs(class_name):
    if db.DB_ENABLED:
        return jsonify(db.get_documents(class_name))
    docs = load_docs()
    return jsonify(docs.get(class_name, []))


@app.route("/api/docs/<class_name>", methods=["POST"])
def post_doc(class_name):
    title = request.form.get("title", "").strip()
    body = request.form.get("body", "").strip()
    doctype = request.form.get("doctype", "درس")
    teacher_name = request.form.get("teacher_name", "")
    subject = request.form.get("subject", "")

    if not title:
        return jsonify({"ok": False, "error": "عنوان الوثيقة مطلوب."}), 400

    file_url = None
    file_name = None
    file_obj = request.files.get("file")
    if file_obj and file_obj.filename:
        safe_name = f"{int(time.time()*1000)}_{file_obj.filename}"
        file_obj.save(os.path.join(UPLOAD_DIR, safe_name))
        file_name = file_obj.filename
        file_url = f"/uploads/{safe_name}"

    if db.DB_ENABLED:
        db.add_document(class_name, title, body, doctype, teacher_name, subject, file_url, file_name)
        return jsonify({"ok": True})

    docs = load_docs()
    docs.setdefault(class_name, [])
    docs[class_name].insert(0, {
        "title": title,
        "body": body,
        "doctype": doctype,
        "teacher": teacher_name,
        "subject": subject,
        "fileUrl": file_url,
        "fileName": file_name,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    save_docs(docs)
    return jsonify({"ok": True})


# ---------------- Student search (autocomplete — used by teacher/supervisor/counselor) ----------------

@app.route("/api/students/search")
def search_students():
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify([])
    results = [
        dict(s) for s in STUDENTS
        if q in s["last_name"] or q in s["first_name"] or q in f"{s['last_name']} {s['first_name']}"
    ]
    phones = db.get_parent_phones_bulk() if db.DB_ENABLED else load_json(PARENT_PHONES_FILE, {})
    for r in results:
        r["parent_phone"] = phones.get(str(r["national_id"]), "")
    return jsonify(results[:15])


# ---------------- Summons (استدعاء تلميذ) ----------------

SUMMONS_FILE = os.path.join(os.path.dirname(DATA_FILE), "summons.json")
REGISTRATIONS_FILE = os.path.join(os.path.dirname(DATA_FILE), "registrations.json")


# ---------------- Registrations (استمارة التسجيل 2026/2027) ----------------

REGISTRATION_REQUIRED_FIELDS = [
    "national_id", "father_name",
    "mother_last_name", "mother_first_name", "address",
    "parent_phone", "parent_whatsapp",
]

LEVEL_LABELS = {
    "1": "السنة الأولى متوسط",
    "2": "السنة الثانية متوسط",
    "3": "السنة الثالثة متوسط",
    "4": "السنة الرابعة متوسط",
}


@app.route("/api/registrations", methods=["POST"])
def create_registration():
    data = request.get_json(force=True)
    cleaned = {}
    missing = []
    for field in REGISTRATION_REQUIRED_FIELDS:
        val = (data.get(field) or "").strip()
        if not val:
            missing.append(field)
        cleaned[field] = val
    if missing:
        return jsonify({"ok": False, "error": "يرجى تعبئة جميع الخانات المطلوبة، واختيار التلميذ(ة) من نتائج البحث."}), 400

    nid = cleaned["national_id"]
    stu = next((s for s in STUDENTS if str(s["national_id"]) == nid), None)
    if not stu:
        return jsonify({
            "ok": False,
            "error": "هذا التلميذ غير موجود ضمن قائمة تلاميذ المؤسسة. التسجيل عبر هذه الاستمارة متاح فقط لتلاميذ المؤسسة الحاليين — يرجى التواصل مع إدارة المؤسسة."
        }), 403

    level = stu["class"][0]  # first digit of class name, e.g. '2م1' -> '2'
    record = {
        "national_id": nid,
        "last_name": stu["last_name"],
        "first_name": stu["first_name"],
        "class_name": stu["class"],
        "level": level,
        "father_name": cleaned["father_name"],
        "mother_last_name": cleaned["mother_last_name"],
        "mother_first_name": cleaned["mother_first_name"],
        "address": cleaned["address"],
        "parent_phone": cleaned["parent_phone"],
        "parent_whatsapp": cleaned["parent_whatsapp"],
    }

    if db.DB_ENABLED:
        try:
            new_id = db.add_registration(record)
        except ValueError:
            return jsonify({"ok": False, "error": "تم تسجيل هذا التلميذ مسبقاً."}), 409
    else:
        regs = load_json(REGISTRATIONS_FILE, [])
        if any(r["national_id"] == nid for r in regs):
            return jsonify({"ok": False, "error": "تم تسجيل هذا التلميذ مسبقاً."}), 409
        new_id = (max([r["id"] for r in regs], default=0)) + 1
        regs.insert(0, {
            "id": new_id, **record, "file_complete": False,
            "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
        save_json(REGISTRATIONS_FILE, regs)

    requires_full_docs = level in ("1", "4")
    return jsonify({
        "ok": True,
        "id": new_id,
        "receipt": {
            "last_name": stu["last_name"],
            "first_name": stu["first_name"],
            "class_name": stu["class"],
            "level_label": LEVEL_LABELS[level],
            "requires_full_docs": requires_full_docs,
        }
    })


@app.route("/api/registrations", methods=["GET"])
def list_registrations():
    if db.DB_ENABLED:
        rows = db.get_registrations()
    else:
        rows = load_json(REGISTRATIONS_FILE, [])
    for r in rows:
        r["level_label"] = LEVEL_LABELS.get(r["level"], r["level"])
    return jsonify(rows)


@app.route("/api/registrations/<int:reg_id>/complete", methods=["POST"])
def mark_registration_complete(reg_id):
    data = request.get_json(force=True)
    complete = bool(data.get("complete"))
    if db.DB_ENABLED:
        db.set_registration_complete(reg_id, complete)
    else:
        regs = load_json(REGISTRATIONS_FILE, [])
        for r in regs:
            if r["id"] == reg_id:
                r["file_complete"] = complete
        save_json(REGISTRATIONS_FILE, regs)
    return jsonify({"ok": True})


@app.route("/api/registrations/summary")
def registrations_summary():
    if db.DB_ENABLED:
        regs = db.get_registrations()
    else:
        regs = load_json(REGISTRATIONS_FILE, [])
    total = len(regs)
    completed = len([r for r in regs if r.get("file_complete")])
    by_level = {}
    for r in regs:
        if r.get("file_complete"):
            by_level[r["level"]] = by_level.get(r["level"], 0) + 1
    return jsonify({
        "total_submitted": total,
        "total_completed": completed,
        "completed_by_level": {LEVEL_LABELS.get(k, k): v for k, v in by_level.items()},
    })


@app.route("/api/summons", methods=["POST"])
def create_summons():
    data = request.get_json(force=True)
    nid = str(data.get("national_id", ""))
    reason = (data.get("reason") or "").strip()
    date = data.get("date") or today_str()
    time_ = data.get("time") or datetime.now().strftime("%H:%M")
    requested_by_role = data.get("requested_by_role") or ""
    requested_by_name = data.get("requested_by_name") or ""
    on_behalf_of = (data.get("on_behalf_of") or "").strip()

    stu = next((s for s in STUDENTS if str(s["national_id"]) == nid), None)
    if not stu:
        return jsonify({"ok": False, "error": "التلميذ غير موجود."}), 404
    if not reason:
        return jsonify({"ok": False, "error": "سبب الاستدعاء مطلوب."}), 400

    if db.DB_ENABLED:
        db.add_summons(nid, stu["last_name"], stu["first_name"], stu["class"],
                        requested_by_role, requested_by_name, on_behalf_of, reason, date, time_)
        return jsonify({"ok": True})

    summons = load_json(SUMMONS_FILE, [])
    summons.insert(0, {
        "national_id": nid, "last_name": stu["last_name"], "first_name": stu["first_name"], "class": stu["class"],
        "requested_by_role": requested_by_role, "requested_by_name": requested_by_name,
        "on_behalf_of": on_behalf_of,
        "reason": reason, "date": date, "time": time_,
    })
    save_json(SUMMONS_FILE, summons)
    return jsonify({"ok": True})


@app.route("/api/summons")
def list_summons():
    date = request.args.get("date") or today_str()
    if db.DB_ENABLED:
        return jsonify(db.get_summons_for_date(date))
    summons = [s for s in load_json(SUMMONS_FILE, []) if s["date"] == date]
    return jsonify(summons)


@app.route("/api/summons/student/<national_id>")
def student_summons(national_id):
    if db.DB_ENABLED:
        return jsonify(db.get_summons_for_student(str(national_id)))
    summons = [s for s in load_json(SUMMONS_FILE, []) if s["national_id"] == str(national_id)]
    return jsonify(summons)


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename, as_attachment=False)


# ---------------- Announcements (أمانة المدير) ----------------

@app.route("/api/announcements", methods=["GET"])
def list_announcements():
    if db.DB_ENABLED:
        return jsonify(db.get_announcements())
    return jsonify(load_json(ANN_FILE, []))


@app.route("/api/announcements", methods=["POST"])
def create_announcement():
    title = (request.form.get("title") or "").strip()
    body = (request.form.get("body") or "").strip()
    target = request.form.get("target") or "all"
    author = request.form.get("author") or "إدارة المؤسسة"
    if not title:
        return jsonify({"ok": False, "error": "عنوان الإعلان مطلوب."}), 400

    file_url = None
    file_name = None
    file_obj = request.files.get("file")
    if file_obj and file_obj.filename:
        safe_name = f"{int(time.time()*1000)}_{file_obj.filename}"
        file_obj.save(os.path.join(UPLOAD_DIR, safe_name))
        file_name = file_obj.filename
        file_url = f"/uploads/{safe_name}"

    if db.DB_ENABLED:
        db.add_announcement(title, body, target, author, file_url, file_name)
        return jsonify({"ok": True})

    anns = load_json(ANN_FILE, [])
    anns.insert(0, {
        "title": title, "body": body, "target": target, "author": author,
        "fileUrl": file_url, "fileName": file_name,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    save_json(ANN_FILE, anns)
    return jsonify({"ok": True})


def _announcement_matches(target, **ctx):
    if target == "all":
        return True
    kind = ctx.get("kind")
    if kind == "student":
        nid, cls = ctx.get("national_id"), ctx.get("class")
        if target == "students:all":
            return True
        if target == f"students:class:{cls}":
            return True
        if target.startswith("students:level:") and cls:
            return cls[0] == target.split(":", 2)[2]
        if target == f"students:one:{nid}":
            return True
    elif kind == "teacher":
        eid = ctx.get("employee_id")
        if target == "teachers:all":
            return True
        if target == f"teachers:one:{eid}":
            return True
    elif kind == "monitor":
        eid = ctx.get("employee_id")
        if target == "monitors:all":
            return True
        if target == f"monitors:one:{eid}":
            return True
    elif kind == "role":
        return target == f"role:{ctx.get('role')}"
    return False


@app.route("/api/announcements/for/students")
def announcements_for_students():
    nid = request.args.get("national_id", "")
    cls = request.args.get("class", "")
    anns = db.get_announcements() if db.DB_ENABLED else load_json(ANN_FILE, [])
    result = [a for a in anns if _announcement_matches(a["target"], kind="student", national_id=nid, **{"class": cls})]
    return jsonify(result)


@app.route("/api/announcements/for/teachers")
def announcements_for_teachers():
    eid = request.args.get("employee_id", "")
    anns = db.get_announcements() if db.DB_ENABLED else load_json(ANN_FILE, [])
    result = [a for a in anns if _announcement_matches(a["target"], kind="teacher", employee_id=eid)]
    return jsonify(result)


@app.route("/api/announcements/for/role/<role>")
def announcements_for_role(role):
    """role = 'supervisor' | 'counselor' | 'bursar'"""
    anns = db.get_announcements() if db.DB_ENABLED else load_json(ANN_FILE, [])
    result = [a for a in anns if _announcement_matches(a["target"], kind="role", role=role)]
    return jsonify(result)


# ---------------- Staff directory ----------------

@app.route("/api/staff/all")
def staff_all():
    return jsonify(STAFF)


@app.route("/api/staff/teaching-and-supervisory")
def staff_teaching_and_supervisory():
    result = [
        s for s in STAFF
        if "أستاذ" in (s.get("role") or "") or "مشرف" in (s.get("role") or "")
    ]
    return jsonify(result)


# ---------------- Absences — students (الناظر) ----------------

@app.route("/api/absences", methods=["POST"])
def mark_absences():
    data = request.get_json(force=True)
    class_name = data.get("class")
    date = data.get("date") or today_str()
    records = data.get("records", [])  # [{national_id, reason, hours}]
    if not class_name:
        return jsonify({"ok": False, "error": "القسم مطلوب."}), 400

    roster = {str(s["national_id"]): s for s in STUDENTS if s["class"] == class_name}
    clean_records = []
    for rec in records:
        nid = str(rec.get("national_id", ""))
        s = roster.get(nid)
        if s:
            clean_records.append({
                "national_id": nid, "last_name": s["last_name"], "first_name": s["first_name"],
                "reason": (rec.get("reason") or "").strip(),
                "hours": rec.get("hours") or 0,
            })

    if db.DB_ENABLED:
        db.save_student_absences(class_name, date, clean_records)
        return jsonify({"ok": True, "count": len(clean_records)})

    absences = load_json(ABS_FILE, [])
    absences = [a for a in absences if not (a["class"] == class_name and a["date"] == date)]
    for rec in clean_records:
        absences.append({**rec, "class": class_name, "date": date})
    save_json(ABS_FILE, absences)
    return jsonify({"ok": True, "count": len(clean_records)})


@app.route("/api/absences")
def get_absences():
    date = request.args.get("date") or today_str()
    class_name = request.args.get("class")
    if db.DB_ENABLED:
        return jsonify(db.get_student_absences(date, class_name))
    absences = load_json(ABS_FILE, [])
    absences = [a for a in absences if a["date"] == date]
    if class_name:
        absences = [a for a in absences if a["class"] == class_name]
    return jsonify(absences)


# ---------------- Absences — staff: teachers & supervisors (الناظر) ----------------

@app.route("/api/staff-absences", methods=["POST"])
def mark_staff_absences():
    data = request.get_json(force=True)
    date = data.get("date") or today_str()
    records = data.get("records", [])  # [{employee_id, reason, hours}]

    staff_by_id = {str(s["employee_id"]): s for s in STAFF}
    clean_records = []
    for rec in records:
        eid = str(rec.get("employee_id", ""))
        s = staff_by_id.get(eid)
        if s:
            clean_records.append({
                "employee_id": eid, "last_name": s["last_name"], "first_name": s["first_name"],
                "role": s["role"],
                "reason": (rec.get("reason") or "").strip(),
                "hours": rec.get("hours") or 0,
            })

    if db.DB_ENABLED:
        db.save_staff_absences(date, clean_records)
        return jsonify({"ok": True, "count": len(clean_records)})

    staff_abs = load_json(STAFF_ABS_FILE, [])
    staff_abs = [a for a in staff_abs if a["date"] != date]
    for rec in clean_records:
        staff_abs.append({**rec, "date": date})
    save_json(STAFF_ABS_FILE, staff_abs)
    return jsonify({"ok": True, "count": len(clean_records)})


@app.route("/api/staff-absences")
def get_staff_absences():
    date = request.args.get("date") or today_str()
    if db.DB_ENABLED:
        return jsonify(db.get_staff_absences(date))
    staff_abs = [a for a in load_json(STAFF_ABS_FILE, []) if a["date"] == date]
    return jsonify(staff_abs)


# ---------------- Guidance (مستشار التوجيه) ----------------

@app.route("/api/guidance/<national_id>", methods=["GET"])
def get_guidance(national_id):
    if db.DB_ENABLED:
        return jsonify(db.get_guidance(str(national_id)))
    guid = load_json(GUID_FILE, {})
    return jsonify(guid.get(str(national_id), []))


@app.route("/api/guidance/<national_id>", methods=["POST"])
def add_guidance(national_id):
    data = request.get_json(force=True)
    reason = (data.get("reason") or "").strip()
    actions = (data.get("actions") or "").strip()
    author = data.get("author") or "مستشار التوجيه"
    if not reason:
        return jsonify({"ok": False, "error": "سبب اللقاء مطلوب."}), 400

    if db.DB_ENABLED:
        db.add_guidance(str(national_id), reason, actions, author)
        return jsonify({"ok": True})

    guid = load_json(GUID_FILE, {})
    guid.setdefault(str(national_id), [])
    guid[str(national_id)].insert(0, {
        "reason": reason, "actions": actions, "author": author,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    save_json(GUID_FILE, guid)
    return jsonify({"ok": True})


@app.route("/api/guidance/recent")
def recent_guidance():
    """All interviews for a given date, across all students — for the director's report
    and for the counselor's own overview list."""
    date = request.args.get("date") or today_str()

    if db.DB_ENABLED:
        rows = db.get_guidance_for_date(date)
        result = []
        for n in rows:
            stu = next((s for s in STUDENTS if str(s["national_id"]) == n["national_id"]), None)
            result.append({
                "national_id": n["national_id"],
                "last_name": stu["last_name"] if stu else "",
                "first_name": stu["first_name"] if stu else n["national_id"],
                "class": stu["class"] if stu else "—",
                "reason": n["reason"], "actions": n["actions"], "author": n["author"], "date": n["date"],
            })
        return jsonify(result)

    guid = load_json(GUID_FILE, {})
    result = []
    for nid, notes in guid.items():
        stu = next((s for s in STUDENTS if str(s["national_id"]) == nid), None)
        for n in notes:
            if n["date"].startswith(date):
                result.append({
                    "national_id": nid,
                    "last_name": stu["last_name"] if stu else "",
                    "first_name": stu["first_name"] if stu else nid,
                    "class": stu["class"] if stu else "—",
                    **n,
                })
    return jsonify(result)


# ---------------- Director's daily report ----------------

@app.route("/api/director/daily-report")
def daily_report():
    date = request.args.get("date") or today_str()

    if db.DB_ENABLED:
        logs = db.get_logins_for_date(date)
        students_logged = [l for l in logs if l["role"] == "student"]
        teachers_logged = [l for l in logs if l["role"] == "teacher"]
        admins_logged = [l for l in logs if l["role"] == "admin"]

        absences = db.get_student_absences(date)
        staff_absences = db.get_staff_absences(date)
        docs_today = db.get_documents_for_date(date)

        rows = db.get_guidance_for_date(date)
        guidance_today = []
        for n in rows:
            stu = next((s for s in STUDENTS if str(s["national_id"]) == n["national_id"]), None)
            guidance_today.append({
                "national_id": n["national_id"],
                "last_name": stu["last_name"] if stu else "",
                "first_name": stu["first_name"] if stu else n["national_id"],
                "class": stu["class"] if stu else "—",
                "reason": n["reason"], "actions": n["actions"], "author": n["author"], "date": n["date"],
            })
    else:
        logs = [l for l in load_json(LOGIN_LOG_FILE, []) if l["date"] == date]
        students_logged = [l for l in logs if l["role"] == "student"]
        teachers_logged = [l for l in logs if l["role"] == "teacher"]
        admins_logged = [l for l in logs if l["role"] == "admin"]

        absences = [a for a in load_json(ABS_FILE, []) if a["date"] == date]
        staff_absences = [a for a in load_json(STAFF_ABS_FILE, []) if a["date"] == date]

        docs = load_docs()
        docs_today = []
        for class_name, items in docs.items():
            for d in items:
                if d["date"].startswith(date):
                    docs_today.append({"class": class_name, **d})

        guid = load_json(GUID_FILE, {})
        guidance_today = []
        for nid, notes in guid.items():
            stu = next((s for s in STUDENTS if str(s["national_id"]) == nid), None)
            for n in notes:
                if n["date"].startswith(date):
                    guidance_today.append({
                        "national_id": nid,
                        "last_name": stu["last_name"] if stu else "",
                        "first_name": stu["first_name"] if stu else nid,
                        "class": stu["class"] if stu else "—",
                        **n,
                    })

    return jsonify({
        "date": date,
        "logins": {
            "students": students_logged,
            "teachers": teachers_logged,
            "admins": admins_logged,
        },
        "absences": absences,
        "staff_absences": staff_absences,
        "documents_sent": docs_today,
        "guidance_interviews": guidance_today,
        "totals": {
            "students_in_school": len(STUDENTS),
            "teachers_in_school": len({a["employee_id"] for a in ASSIGNMENTS}),
        }
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
